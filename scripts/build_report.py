"""
build_report.py  –  Generate Markdown + XLSX drift report from drift.json.

Usage:
    python scripts/build_report.py <out/drift.json> <out/DRIFT_REPORT.md> <out/drift.xlsx>

Requires: openpyxl
"""

import json
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Verdict ordering for the report
VERDICT_ORDER = ["DRIFTED", "UNDOCUMENTED", "UNVERIFIABLE", "CONFIRMED",
                 "MISSING_VERDICT"]

VERDICT_COLOURS = {
    "DRIFTED":         "FFC7CE",   # pale red
    "UNDOCUMENTED":    "FFEB9C",   # pale orange
    "UNVERIFIABLE":    "DDEBF7",   # pale blue
    "CONFIRMED":       "C6EFCE",   # pale green
    "MISSING_VERDICT": "F2F2F2",   # grey
}

# ── markdown helpers ──────────────────────────────────────────────────────────

def _citations_str(citations: list[dict]) -> str:
    if not citations:
        return "—"
    parts = []
    for c in citations:
        path = c.get("file", "?")
        sl   = c.get("start_line")
        el   = c.get("end_line")
        if sl and el:
            parts.append(f"`{path}:{sl}-{el}`")
        elif sl:
            parts.append(f"`{path}:{sl}`")
        else:
            parts.append(f"`{path}`")
    return ", ".join(parts)


def _build_markdown(drift: dict) -> str:
    lines = []
    summary = drift["summary"]

    lines.append("# Drift Report\n")

    # ── summary table ──
    lines.append("## Summary\n")
    lines.append("| Metric | Count |")
    lines.append("|---|---|")
    lines.append(f"| Total rules in spec | {summary['total_rules']} |")
    lines.append(f"| Verdicts rendered | {summary['total_verdicts']} |")
    lines.append(f"| Undocumented behaviours | {summary['total_undocumented']} |")
    lines.append(f"| Missing verdicts | {summary['total_missing']} |")
    lines.append(f"| Validation problems | {summary['validation_problems']} |")
    lines.append("")
    lines.append("### By verdict")
    lines.append("| Verdict | Count |")
    lines.append("|---|---|")
    for verdict, count in summary.get("by_verdict", {}).items():
        lines.append(f"| {verdict} | {count} |")
    lines.append("")
    lines.append("### By confidence")
    lines.append("| Confidence | Count |")
    lines.append("|---|---|")
    for conf, count in summary.get("by_confidence", {}).items():
        lines.append(f"| {conf} | {count} |")
    lines.append("")

    # bucket verdicts
    buckets: dict[str, list[dict]] = {v: [] for v in VERDICT_ORDER}
    for v in drift.get("verdicts", []):
        verdict = v.get("verdict", "?")
        if verdict in buckets:
            buckets[verdict].append(v)
        else:
            buckets.setdefault(verdict, []).append(v)

    for u in drift.get("undocumented", []):
        buckets["UNDOCUMENTED"].append(u)

    for m in drift.get("missing", []):
        buckets["MISSING_VERDICT"].append(m)

    # ── DRIFTED ──
    if buckets["DRIFTED"]:
        lines.append("---\n## Drifted Rules\n")
        for v in buckets["DRIFTED"]:
            lines.append(f"### {v.get('rule_id')} — {v.get('confidence','')} confidence")
            lines.append(f"- **Spec says:** {v.get('spec_says','—')}")
            lines.append(f"- **Code does:** {v.get('code_does','—')}")
            lines.append(f"- **Citations:** {_citations_str(v.get('citations',[]))}")
            if v.get("note"):
                lines.append(f"- **Note:** {v['note']}")
            lines.append("")

    # ── UNDOCUMENTED ──
    if buckets["UNDOCUMENTED"]:
        lines.append("---\n## Undocumented Behaviours\n")
        for u in buckets["UNDOCUMENTED"]:
            lines.append(f"### {u.get('title','(no title)')} — {u.get('confidence','')} confidence")
            lines.append(f"- **Code does:** {u.get('code_does','—')}")
            lines.append(f"- **Why it matters:** {u.get('why_it_matters','—')}")
            lines.append(f"- **Citations:** {_citations_str(u.get('citations',[]))}")
            lines.append("")

    # ── UNVERIFIABLE ──
    if buckets["UNVERIFIABLE"]:
        lines.append("---\n## Unverifiable Rules\n")
        for v in buckets["UNVERIFIABLE"]:
            lines.append(f"### {v.get('rule_id')} — {v.get('confidence','')} confidence")
            lines.append(f"- **Spec says:** {v.get('spec_says','—')}")
            lines.append(f"- **Code does:** {v.get('code_does','—')}")
            if v.get("note"):
                lines.append(f"- **Note:** {v['note']}")
            lines.append("")

    # ── CONFIRMED (compact table) ──
    if buckets["CONFIRMED"]:
        lines.append("---\n## Confirmed Rules (compact)\n")
        lines.append("| Rule ID | Confidence | Citations | Note |")
        lines.append("|---|---|---|---|")
        for v in buckets["CONFIRMED"]:
            rid   = v.get("rule_id","")
            conf  = v.get("confidence","")
            cits  = _citations_str(v.get("citations",[]))
            note  = v.get("note","") or ""
            lines.append(f"| {rid} | {conf} | {cits} | {note} |")
        lines.append("")

    # ── MISSING_VERDICT ──
    if buckets["MISSING_VERDICT"]:
        lines.append("---\n## Missing Verdicts\n")
        lines.append("| Rule ID | Section |")
        lines.append("|---|---|")
        for m in buckets["MISSING_VERDICT"]:
            lines.append(f"| {m.get('rule_id','')} | {m.get('section','')} |")
        lines.append("")

    return "\n".join(lines)


# ── xlsx helpers ──────────────────────────────────────────────────────────────

def _header_row(ws, headers: list[str], fill_hex: str = "4472C4"):
    fill   = PatternFill(fill_type="solid", fgColor=fill_hex)
    font   = Font(bold=True, color="FFFFFF")
    for col, header in enumerate(headers, 1):
        cell            = ws.cell(row=1, column=col, value=header)
        cell.fill       = fill
        cell.font       = font
        cell.alignment  = Alignment(wrap_text=True, vertical="top")


def _auto_width(ws, max_width: int = 60):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        width = max(
            (len(str(cell.value)) if cell.value else 0) for cell in col
        )
        ws.column_dimensions[col_letter].width = min(max(width + 4, 10), max_width)


def _add_verdict_sheet(wb: openpyxl.Workbook, sheet_name: str,
                       rows: list[list], headers: list[str], fill_hex: str):
    ws = wb.create_sheet(title=sheet_name)
    _header_row(ws, headers, fill_hex)
    row_fill = PatternFill(fill_type="solid", fgColor=VERDICT_COLOURS.get(sheet_name.upper(), "FFFFFF"))
    for r_idx, row in enumerate(rows, 2):
        for c_idx, val in enumerate(row, 1):
            cell           = ws.cell(row=r_idx, column=c_idx, value=str(val) if val is not None else "")
            cell.fill      = row_fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    _auto_width(ws)
    return ws


def _build_xlsx(drift: dict, xlsx_path: Path):
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    summary = drift["summary"]

    # ── Summary sheet ──
    ws_sum = wb.create_sheet(title="Summary")
    _header_row(ws_sum, ["Metric", "Count"])
    summary_rows = [
        ["Total rules in spec",      summary["total_rules"]],
        ["Verdicts rendered",         summary["total_verdicts"]],
        ["Undocumented behaviours",   summary["total_undocumented"]],
        ["Missing verdicts",          summary["total_missing"]],
        ["Validation problems",       summary["validation_problems"]],
    ]
    for v, c in summary.get("by_verdict", {}).items():
        summary_rows.append([f"Verdict: {v}", c])
    for conf, c in summary.get("by_confidence", {}).items():
        summary_rows.append([f"Confidence: {conf}", c])

    for r_idx, row in enumerate(summary_rows, 2):
        ws_sum.cell(row=r_idx, column=1, value=row[0])
        ws_sum.cell(row=r_idx, column=2, value=row[1])
    _auto_width(ws_sum)

    # bucket by verdict
    buckets: dict[str, list[dict]] = {v: [] for v in VERDICT_ORDER}
    for v in drift.get("verdicts", []):
        verdict = v.get("verdict","?")
        if verdict in buckets:
            buckets[verdict].append(v)

    for u in drift.get("undocumented", []):
        buckets["UNDOCUMENTED"].append(u)

    for m in drift.get("missing", []):
        buckets["MISSING_VERDICT"].append(m)

    # ── DRIFTED sheet ──
    drifted_rows = []
    for v in buckets["DRIFTED"]:
        drifted_rows.append([
            v.get("rule_id",""),
            v.get("confidence",""),
            v.get("spec_says",""),
            v.get("code_does",""),
            _citations_str(v.get("citations",[])),
            v.get("note","") or "",
        ])
    if drifted_rows:
        _add_verdict_sheet(wb, "DRIFTED", drifted_rows,
                           ["Rule ID","Confidence","Spec Says","Code Does","Citations","Note"],
                           VERDICT_COLOURS["DRIFTED"][:6])

    # ── UNDOCUMENTED sheet ──
    undoc_rows = []
    for u in buckets["UNDOCUMENTED"]:
        undoc_rows.append([
            u.get("title",""),
            u.get("confidence",""),
            u.get("code_does",""),
            u.get("why_it_matters",""),
            _citations_str(u.get("citations",[])),
        ])
    if undoc_rows:
        _add_verdict_sheet(wb, "UNDOCUMENTED", undoc_rows,
                           ["Title","Confidence","Code Does","Why It Matters","Citations"],
                           VERDICT_COLOURS["UNDOCUMENTED"][:6])

    # ── UNVERIFIABLE sheet ──
    unver_rows = []
    for v in buckets["UNVERIFIABLE"]:
        unver_rows.append([
            v.get("rule_id",""),
            v.get("confidence",""),
            v.get("spec_says",""),
            v.get("code_does",""),
            v.get("note","") or "",
        ])
    if unver_rows:
        _add_verdict_sheet(wb, "UNVERIFIABLE", unver_rows,
                           ["Rule ID","Confidence","Spec Says","Code Does","Note"],
                           VERDICT_COLOURS["UNVERIFIABLE"][:6])

    # ── CONFIRMED sheet ──
    confirmed_rows = []
    for v in buckets["CONFIRMED"]:
        confirmed_rows.append([
            v.get("rule_id",""),
            v.get("confidence",""),
            _citations_str(v.get("citations",[])),
            v.get("note","") or "",
        ])
    if confirmed_rows:
        _add_verdict_sheet(wb, "CONFIRMED", confirmed_rows,
                           ["Rule ID","Confidence","Citations","Note"],
                           VERDICT_COLOURS["CONFIRMED"][:6])

    # ── MISSING sheet ──
    missing_rows = []
    for m in buckets["MISSING_VERDICT"]:
        missing_rows.append([
            m.get("rule_id",""),
            m.get("section",""),
            m.get("text","")[:120],
        ])
    if missing_rows:
        _add_verdict_sheet(wb, "MISSING_VERDICT", missing_rows,
                           ["Rule ID","Section","Text (truncated)"],
                           VERDICT_COLOURS["MISSING_VERDICT"][:6])

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)


# ── main ───────────────────────────────────────────────────────────────────────

def build(drift_path: Path, md_path: Path, xlsx_path: Path):
    drift = json.loads(drift_path.read_text(encoding="utf-8"))

    md_text = _build_markdown(drift)
    md_path.parent.mkdir(parents=True, exist_ok=True)
    md_path.write_text(md_text, encoding="utf-8")
    print(f"Written: {md_path}  ({len(md_text):,} chars)")

    _build_xlsx(drift, xlsx_path)
    print(f"Written: {xlsx_path}  ({xlsx_path.stat().st_size:,} bytes)")

    # Print the markdown report to stdout
    print("\n" + "="*60)
    print("DRIFT REPORT (Markdown)")
    print("="*60)
    print(md_text)


if __name__ == "__main__":
    if len(sys.argv) != 4:
        print(
            "Usage: python scripts/build_report.py "
            "<out/drift.json> <out/DRIFT_REPORT.md> <out/drift.xlsx>"
        )
        sys.exit(1)
    build(Path(sys.argv[1]), Path(sys.argv[2]), Path(sys.argv[3]))
